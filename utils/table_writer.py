import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from .config import PLOT_OUTPUT_DIR

def write_excel_style_summary_table(df, filename="model_accuracy_summary.xlsx"):
    filepath = os.path.join(PLOT_OUTPUT_DIR, filename)
    wb = Workbook()
    wb.remove(wb.active)  # remove the default sheet

    for model in df["Model"].unique():
        model_df = df[df["Model"] == model]
        table = model_df.pivot(index="QueryLanguage", columns="SchemaType", values="Mean")

        # Convert to percentage format (omit symbol)
        table *= 100
        table = table.round(1)

        # Add row-wise average
        table["avg"] = table.mean(axis=1).round(1)

        # Add column-wise average
        avg_row = table.mean(axis=0).round(1)
        avg_row.name = "avg"
        table = pd.concat([table, pd.DataFrame([avg_row])])

        ws = wb.create_sheet(title=model[:31])  # Excel sheet name max length = 31

        rows = list(dataframe_to_rows(table.reset_index(), index=False, header=True))

        # Replace top-left header cell with model name
        rows[0][0] = model

        # Styles
        center = Alignment(horizontal='center', vertical='center')
        bold = Font(bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write table
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.alignment = center
                cell.border = thin_border
                if r_idx == 1:
                    cell.font = bold

        # Highlight max values per column (excluding 'avg' row)
        schema_cols = table.columns.drop("avg")
        for col_idx, schema in enumerate(schema_cols, start=2):
            col_vals = table.loc[table.index != "avg", schema]
            max_val = col_vals.max()
            for row_idx, val in enumerate(col_vals, start=2):
                if val == max_val:
                    ws.cell(row=row_idx, column=col_idx).font = Font(bold=True)

    wb.save(filepath)
    print(f"✅ Saved Excel summary to {filepath}")
